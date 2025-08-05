import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image
import os
import sys
from datetime import date


def extract_data_final(file_path):
    """Extracts text data. Sets Product Name same as Product Code."""
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    order_id = None
    for row in ws.iter_rows(max_row=10):
        for cell in row:
            if isinstance(cell.value, str):
                value = cell.value.replace('：', ':')
                if "销售单号" in value:
                    try:
                        order_id = value.split(':', 1)[1].strip()
                    except IndexError:
                        pass
    table_header_row_num = None
    for cell in ws['A'] + ws['B'] + ws['C']:
        if isinstance(cell.value, str) and cell.value.strip() == "序号":
            table_header_row_num = cell.row
            break
    if not order_id or not table_header_row_num:
        raise ValueError("Could not find '销售单号' or the table header '序号'.")

    items_df = pd.read_excel(file_path, header=table_header_row_num - 1, engine='openpyxl')
    items_list = []
    for _, row in items_df.iterrows():
        if pd.isna(row.get('序号')):
            break
        part_full_string = str(row['图号']).replace('\n', ' ').strip()
        material_val = str(row['材料'])
        if '规格' in items_df.columns and pd.notna(row['规格']):
            material_val += f" ({row['规格']})"
        items_list.append({
            "seq": int(row['序号']),
            "product_code": part_full_string,
            "product_name": part_full_string,
            "material": material_val,
            "quantity": int(row['数量']),
            "image_filename_base": part_full_string,
        })
    return {"order_id": order_id, "items": items_list}


def create_final_delivery_note(source_data, customer_info, output_filename, script_dir):
    """Creates a minimalistic Delivery Note Excel."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "交货单"

    light_gray = PatternFill(start_color="F5F5F7", end_color="F5F5F7", fill_type="solid")
    company_font = Font(name='PingFang SC', size=20, bold=False, color="1D1D1F")
    title_font = Font(name='PingFang SC', size=32, bold=False, color="1D1D1F")
    label_font = Font(name='PingFang SC', size=12, bold=False, color="86868B")
    value_font = Font(name='PingFang SC', size=12, bold=False, color="1D1D1F")
    table_header_font = Font(name='PingFang SC', size=11, bold=True, color="1D1D1F")

    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center')

    thin_border = Border(bottom=Side(style='thin', color='E5E5EA'))
    table_border = Border(
        left=Side(style='thin', color='E5E5EA'),
        right=Side(style='thin', color='E5E5EA'),
        top=Side(style='thin', color='E5E5EA'),
        bottom=Side(style='thin', color='E5E5EA'),
    )

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 28
    ws.column_dimensions['D'].width = 28
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 15

    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 45
    ws.row_dimensions[3].height = 20
    ws.row_dimensions[4].height = 25
    ws.row_dimensions[5].height = 25
    ws.row_dimensions[6].height = 25
    ws.row_dimensions[7].height = 25
    ws.row_dimensions[8].height = 15

    ws.merge_cells('A1:G1')
    ws['A1'] = "杭州越侬模型科技有限公司"
    ws['A1'].font = company_font
    ws['A1'].alignment = center_align

    ws.merge_cells('A2:G2')
    ws['A2'] = "交货单"
    ws['A2'].font = title_font
    ws['A2'].alignment = center_align

    ws['A4'] = "交货单号"
    ws['A4'].font = label_font
    ws['A4'].alignment = left_align
    ws.merge_cells('B4:C4')
    ws['B4'] = source_data.get('order_id', 'N/A')
    ws['B4'].font = value_font
    ws['B4'].alignment = left_align

    ws['A5'] = "客户名称"
    ws['A5'].font = label_font
    ws['A5'].alignment = left_align
    ws.merge_cells('B5:C5')
    ws['B5'] = customer_info['customer_name']
    ws['B5'].font = value_font
    ws['B5'].alignment = left_align

    ws['A6'] = "联系人"
    ws['A6'].font = label_font
    ws['A6'].alignment = left_align
    ws.merge_cells('B6:C6')
    ws['B6'] = customer_info['contact_person']
    ws['B6'].font = value_font
    ws['B6'].alignment = left_align

    ws['A7'] = "料号"
    ws['A7'].font = label_font
    ws['A7'].alignment = left_align
    ws.merge_cells('B7:C7')
    ws['B7'] = customer_info.get('material_number', '')
    ws['B7'].font = value_font
    ws['B7'].alignment = left_align

    ws['E4'] = "送货日期"
    ws['E4'].font = label_font
    ws['E4'].alignment = left_align
    ws.merge_cells('F4:G4')
    ws['F4'] = date.today().strftime('%Y-%m-%d')
    ws['F4'].font = value_font
    ws['F4'].alignment = left_align

    ws['E5'] = "制单人"
    ws['E5'].font = label_font
    ws['E5'].alignment = left_align
    ws.merge_cells('F5:G5')
    ws['F5'] = customer_info['prepared_by']
    ws['F5'].font = value_font
    ws['F5'].alignment = left_align

    ws['E6'] = "货品总数"
    ws['E6'].font = label_font
    ws['E6'].alignment = left_align
    ws.merge_cells('F6:G6')
    ws['F6'] = sum(item['quantity'] for item in source_data['items'])
    ws['F6'].font = value_font
    ws['F6'].alignment = left_align

    for col in range(1, 8):
        ws.cell(row=8, column=col).border = thin_border

    table_start_row = 9
    headers = ["序号", "产品图片", "产品编号", "产品名称", "材料", "交货数量", "备注"]
    for col, text in enumerate(headers, 1):
        cell = ws.cell(row=table_start_row, column=col, value=text)
        cell.font = table_header_font
        cell.alignment = center_align
        cell.fill = light_gray
        cell.border = table_border

    current_row = table_start_row + 1
    for item in source_data['items']:
        ws.row_dimensions[current_row].height = 60
        ws.cell(row=current_row, column=1, value=item['seq'])
        ws.cell(row=current_row, column=3, value=item['product_code'])
        ws.cell(row=current_row, column=4, value=item['product_name'])
        ws.cell(row=current_row, column=5, value=item['material'])
        ws.cell(row=current_row, column=6, value=item['quantity'])
        for col in range(1, 8):
            cell = ws.cell(row=current_row, column=col)
            cell.alignment = center_align
            cell.border = table_border
            cell.font = Font(name='PingFang SC', size=11, color="1D1D1F")
        base_name = item['image_filename_base']
        possible_ext = ['.png', '.jpg', '.jpeg', '.bmp', '.gif']
        image_path = next(
            (
                os.path.join(script_dir, base_name + ext)
                for ext in possible_ext
                if os.path.exists(os.path.join(script_dir, base_name + ext))
            ),
            None,
        )
        if image_path:
            img = Image(image_path)
            img.height = 75
            img.width = 75
            ws.add_image(img, f'B{current_row}')
        else:
            ws.cell(row=current_row, column=2, value="—").alignment = center_align
        current_row += 1

    ws.row_dimensions[current_row + 2].height = 30
    ws.cell(row=current_row + 2, column=5, value="签收人").font = label_font
    ws.merge_cells(f'F{current_row + 2}:G{current_row + 2}')
    ws.cell(row=current_row + 2, column=6).border = Border(bottom=Side(style='thin', color='E5E5EA'))

    wb.save(output_filename)
    print(f"✅ Successfully created '{output_filename}'")


def find_production_file(folder):
    for name in os.listdir(folder):
        if name.startswith('生产单') and name.lower().endswith(('.xls', '.xlsx')):
            return os.path.join(folder, name)
    return None


def main():
    if len(sys.argv) < 2:
        print("Usage: python generate_delivery_note.py <folder> [customer] [contact] [material] [prepared_by]")
        return 1
    folder = sys.argv[1]
    customer = sys.argv[2] if len(sys.argv) > 2 else ""
    contact = sys.argv[3] if len(sys.argv) > 3 else ""
    material = sys.argv[4] if len(sys.argv) > 4 else ""
    prepared = sys.argv[5] if len(sys.argv) > 5 else ""

    prod_file = find_production_file(folder)
    if not prod_file:
        print("❌ Error: 生产单 not found in folder")
        return 1
    try:
        production_data = extract_data_final(prod_file)
        output_filepath = os.path.join(folder, f"出货单_{production_data['order_id']}.xlsx")
        customer_info = {
            "customer_name": customer,
            "contact_person": contact,
            "material_number": material,
            "prepared_by": prepared,
        }
        create_final_delivery_note(production_data, customer_info, output_filepath, folder)
    except Exception as e:
        print(f"❌ PROCESSING ERROR: {e}")
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
